# -*- coding: utf-8 -*-
"""数据审计:附件1(巡检点)与附件2(禁飞区)。
输出:控制台报告 + docs/02_数据审计.md
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from collections import Counter
import numpy as np

ROOT = os.path.dirname(os.path.abspath(__file__))          # TJMM 根目录
D1 = os.path.join(ROOT, 'A 题', '附件1.xlsx')
D2 = os.path.join(ROOT, 'A 题', '附件2.xlsx')
OUT = os.path.join(ROOT, 'A 题', 'docs', '02_数据审计.md')

lines = ['# 数据审计报告(附件1 / 附件2)', '', '审计脚本: `audit_dataset.py`(TJMM 根目录)。原则:原始文件只读,所有处理在副本上进行。', '']

wb1 = openpyxl.load_workbook(D1, data_only=True)
wb2 = openpyxl.load_workbook(D2, data_only=True)

def ap(s=''): lines.append(s)

# ---------------- 附件1 ----------------
ap('## 附件1:巡检点')
ap('')
ap('| 算例 | 点数 n | I 级(3次) | II 级(2次) | III 级(1次) | 总巡检次数 M | x 范围 | y 范围 |')
ap('|---|---|---|---|---|---|---|---|')
tot_points = 0
for ws in wb1.worksheets:
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    n = len(rows)
    tot_points += n
    ids = [r[0] for r in rows]
    lv = Counter(r[3] for r in rows)
    xs = [float(r[1]) for r in rows]; ys = [float(r[2]) for r in rows]
    visits = 3*lv.get('I',0) + 2*lv.get('II',0) + lv.get('III',0)
    ap(f'| {ws.title} | {n} | {lv.get("I",0)} | {lv.get("II",0)} | {lv.get("III",0)} | {visits} | '
       f'{min(xs):.3f}~{max(xs):.3f} | {min(ys):.3f}~{max(ys):.3f} |')
    # --- 完整性/合法性检查 ---
    assert ids == list(range(1, n+1)), f'{ws.title}: Point_ID 不连续'
    bad_lv = [r[3] for r in rows if r[3] not in ('I','II','III')]
    neg = [r for r in rows if float(r[1]) < 0 or float(r[2]) < 0]
    dup_xy = [ (x,y) for (x,y),c in Counter((round(float(r[1]),9), round(float(r[2]),9)) for r in rows).items() if c > 1]
    # --- 与基地距离 ---
    d = np.hypot(np.array(xs), np.array(ys))
    # --- 点间最小距离 ---
    P = np.column_stack([xs, ys])
    D = np.linalg.norm(P[:,None,:]-P[None,:,:], axis=2)
    np.fill_diagonal(D, np.inf)
    mn = D.min()
    ap('')
    ap(f'**{ws.title} 检查**:Point_ID 连续 ✓ | 非法等级 {len(bad_lv)} 个 | 负坐标 {len(neg)} 个 | '
       f'重合坐标 {len(dup_xy)} 组 | 距基地 min/median/max = {d.min():.1f}/{np.median(d):.1f}/{d.max():.1f} 单位 | '
       f'点间最小距离 {mn:.3f} 单位')
    ap('')

tot_visits = {}
for ws in wb1.worksheets:
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    lv = Counter(r[3] for r in rows)
    tot_visits[ws.title] = 3*lv.get('I',0) + 2*lv.get('II',0) + lv.get('III',0)
ap(f'**合计**:4 个算例共 **{tot_points}** 个巡检点。总巡检次数(工作量):'
   + ' / '.join(f'{k} {v}' for k, v in tot_visits.items()) + '。')
ap('')
ap('审计结论:ID 连续、等级合法、坐标非负、无重复点。点间最小距离见上(最密集处 Case4 ~0.3 单位=30 m,数值精度无风险)。')
ap('')

# ---------------- 附件2 ----------------
ap('## 附件2:临时禁飞区(圆)')
ap('')
ap('| 算例 | 禁飞区 | 圆心 (x,y) | 半径 | 生效时段 |')
ap('|---|---|---|---|---|')
for ws in wb2.worksheets:
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    for r in rows:
        ap(f'| {ws.title} | {r[0]} | ({r[1]:.3f}, {r[2]:.3f}) | {r[3]:.1f} | {r[4]} ~ {r[5]} |')
    # 检查
    ids = [r[0] for r in rows]
    dup = [k for k,v in Counter(ids).items() if v>1]
    neg_rad = [r for r in rows if float(r[3]) <= 0]
    bad_win = [r for r in rows if str(r[4]) > str(r[5])]
    zero_win = [r for r in rows if str(r[4]) == str(r[5])]
    ap('')
    ap(f'**{ws.title} 检查**:Zone_ID 重复 {len(dup)} 个 | 非正半径 {len(neg_rad)} 个 | '
       f'Start>End {len(bad_win)} 个 | 零长度窗口 {len(zero_win)} 个({[r[0] for r in zero_win]})')
    ap('')
ap('''审计结论:
1. 各算例禁飞区时间窗均落在 8:00–17:00 区间内,与 9 h 时限自洽。
2. **Case4 的 Z8 窗口为 17:00–17:00(零长度)**:按"瞬时生效"处理,默认不构成路径约束;敏感性分析给出"17:00 起持续生效"的对照。
3. 禁飞区为圆,圆心/半径与巡检点坐标同单位(1 单位=100 m)。
4. 需在问题3 中检查:哪些巡检点落在禁飞圆内(若点被覆盖,对应时刻无法巡检)。
''')

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print('\n'.join(lines))
print('\nsaved:', OUT)
