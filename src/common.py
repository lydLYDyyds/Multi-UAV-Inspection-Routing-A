# -*- coding: utf-8 -*-
"""公共模块:常量、数据加载、时间/距离换算。

单位约定:
- 坐标单位 u:1 u = 100 m = 0.1 km
- 速度 v = 55 km/h -> 单位距离飞行时间 τ = 0.1/55 h = 6/55 min ≈ 0.109091 min/u
- 单次巡检服务时间 s = 5 min
"""
import os
import openpyxl
import numpy as np

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # TJMM 根目录
A_DIR = os.path.join(ROOT, 'A 题')

V_KMH = 55.0
UNIT_KM = 0.1
SERVICE_MIN = 5.0
TAU_MIN = UNIT_KM / V_KMH * 60.0            # min per unit distance
HORIZON_MIN = 9 * 60.0                      # 9 h
START_MIN = 0.0                             # 8:00 起算,相对时间
LEVEL_VISITS = {'I': 3, 'II': 2, 'III': 1}


def load_points(case):
    """附件1 某算例 -> dict(ids, xy, level, visits)"""
    wb = openpyxl.load_workbook(os.path.join(A_DIR, '附件1.xlsx'), data_only=True)
    ws = wb[case]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    xy = np.array([[float(r[1]), float(r[2])] for r in rows])
    ids = np.array([int(r[0]) for r in rows], dtype=int)
    level = np.array([str(r[3]) for r in rows])
    visits = np.array([LEVEL_VISITS[l] for l in level], dtype=int)
    return dict(ids=ids, xy=xy, level=level, visits=visits)


def load_zones(case):
    """附件2 某算例 -> list of dict(zone_id, c, r, t0, t1) 时间以 8:00 为 0 的分钟数"""
    wb = openpyxl.load_workbook(os.path.join(A_DIR, '附件2.xlsx'), data_only=True)
    ws = wb[case]
    zones = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        def to_min(s):
            h, m = str(s).split(':')
            return (int(h) - 8) * 60 + int(m)
        zones.append(dict(zone_id=str(row[0]), c=np.array([float(row[1]), float(row[2])]),
                          r=float(row[3]), t0=to_min(row[4]), t1=to_min(row[5])))
    return zones


def dist(a, b):
    """欧氏距离(坐标单位)"""
    return float(np.hypot(*(np.asarray(a, float) - np.asarray(b, float))))


def travel_min(d_units):
    """距离(单位)->飞行时间(分钟)"""
    return d_units * TAU_MIN


if __name__ == '__main__':
    for c in ['Case1', 'Case2', 'Case3', 'Case4']:
        p = load_points(c)
        z = load_zones(c)
        print(c, 'points', p['xy'].shape, 'visits', p['visits'].sum(),
              'zones', len(z), 'TAU_MIN=%.5f' % TAU_MIN)
