# -*- coding: utf-8 -*-
"""问题3 结果独立校验 + 导出 result3.xlsx(表4 + 调度明细含航路点/等待)。"""
import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import numpy as np
import openpyxl
from common import load_points, load_zones, A_DIR

CASES = ['Case1', 'Case2', 'Case3', 'Case4']
TAU_MIN = 0.1 / 55.0 * 60.0
SERVICE_MIN = 5.0


def hhmm(minutes):
    return f'{int(8 + minutes // 60):02d}:{int(minutes % 60):02d}'


def path_len(pts):
    pts = [np.asarray(q, float) for q in pts]
    return sum(float(np.linalg.norm(u - v)) for u, v in zip(pts[:-1], pts[1:]))


def seg_hits(a, b, z):
    a = np.asarray(a, float); b = np.asarray(b, float); c = z['c']; r = z['r']
    ab = b - a; L2 = float(ab @ ab)
    if L2 == 0:
        return float(np.linalg.norm(a - c)) < r - 1e-9
    tt = float(np.clip((c - a) @ ab / L2, 0, 1))
    return float(np.linalg.norm(a + tt * ab - c)) < r - 1e-9


def seg_cross_params(a, b, z):
    """解析求交(与求解器独立实现):返回航段在圆内的参数区间 (lo,hi) 或 None。
    aλ²+bλ+c=0 判交;相切/端点接触不算(开圆约定)。"""
    a = np.asarray(a, float); b = np.asarray(b, float)
    d = b - a; c = z['c']; r = z['r']
    A = float(d @ d)
    if A <= 1e-12:
        return (0.0, 1.0) if float(np.linalg.norm(a - c)) < r - 1e-9 else None
    B = 2.0 * float((a - c) @ d)
    C = float((a - c) @ (a - c)) - r * r
    disc = B * B - 4.0 * A * C
    if disc <= 1e-12 * max(1.0, A * A):
        return None
    sq = np.sqrt(disc)
    l1, l2 = (-B - sq) / (2.0 * A), (-B + sq) / (2.0 * A)
    lo, hi = max(0.0, l1), min(1.0, l2)
    if lo > hi + 1e-12 or hi - lo <= 1e-7:
        return None
    return (lo, hi)


def leg_cross_windows(leg, t_dep, t_arr, z):
    """折线 leg 在 [t_dep,t_arr] 内穿越区 z 圆内部的时间区间列表(分钟,8:00 起算)。"""
    total = sum(float(np.linalg.norm(np.asarray(u, float) - np.asarray(v, float)))
                for u, v in zip(leg[:-1], leg[1:]))
    if total <= 1e-12:
        return []
    dur = t_arr - t_dep
    out, cum = [], 0.0
    for u, v in zip(leg[:-1], leg[1:]):
        seg_len = float(np.linalg.norm(np.asarray(u, float) - np.asarray(v, float)))
        pr = seg_cross_params(u, v, z)
        if pr is not None:
            lo, hi = pr
            out.append((t_dep + (cum + lo * seg_len) / total * dur,
                        t_dep + (cum + hi * seg_len) / total * dur))
        cum += seg_len
    return out


def poly_hits(pts, z):
    return any(seg_hits(u, v, z) for u, v in zip(pts[:-1], pts[1:]))


def seg_times(leg, t_dep, leg_waits):
    """各子段真实飞行时段:pos≥1 的悬停插在对应航路点之后。"""
    cum = float(t_dep)
    times = []
    for i, (u, v) in enumerate(zip(leg[:-1], leg[1:])):
        if i >= 1:
            cum += sum(wt for (t_s, wt, pos) in leg_waits if pos == i)
        L = float(np.linalg.norm(np.asarray(u, float) - np.asarray(v, float)))
        times.append((cum, cum + L * TAU_MIN))
        cum += L * TAU_MIN
    return times


def verify(case, routes):
    p = load_points(case)
    xy, visits = p['xy'], p['visits']
    zones = load_zones(case)
    M = int(visits.sum())
    seen = set()
    viol = []
    for r in routes:
        t = 0.0
        for k, (leg, s) in enumerate(zip(r['legs_wp'][:-1], r['seq'])):
            t_dep = r['leg_dep_min'][k]
            t_arr = r['arr_min'][k]
            lw = [(t_s, wt, pos) for (k_, t_s, wt, kind, pos) in r.get('waits', [])
                  if k_ == k + 1]
            exp_arr = t_dep + path_len(leg) * TAU_MIN + \
                sum(wt for (t_s, wt, pos) in lw if pos > 0)
            assert abs(t_arr - exp_arr) < 0.03, \
                f'{case} 机{r["veh"]} 第{k+1}段时刻不一致'
            for (u, v), (t0, t1) in zip(zip(leg[:-1], leg[1:]), seg_times(leg, t_dep, lw)):
                for z in zones:
                    pr = seg_cross_params(u, v, z)
                    if pr is not None:
                        ta = t0 + pr[0] * (t1 - t0)
                        tb = t0 + pr[1] * (t1 - t0)
                        if max(ta, z['t0']) < min(tb, z['t1']):
                            viol.append((r['veh'], 'leg', k + 1, z['zone_id']))
            key = (s['point_id'], s['visit_no'])
            assert key not in seen, f'{case}: 任务 {key} 重复'
            seen.add(key)
            pxy = xy[s['point_id'] - 1]
            for z in zones:
                if float(np.linalg.norm(pxy - z['c'])) < z['r'] - 1e-9 \
                        and max(r['dep_min'][k] - SERVICE_MIN, z['t0']) < min(r['dep_min'][k], z['t1']):
                    viol.append((r['veh'], 'service', k + 1, z['zone_id']))
            assert r['dep_min'][k] >= t_arr + SERVICE_MIN - 0.03, \
                f'{case} 机{r["veh"]} 第{k+1}段离开时刻不符(服务未满5min)'
            # 等待状态硬约束:等待位置位于生效区内即违例
            for (k_, t_s, wt, kind, pos) in r.get('waits', []):
                if k_ != k + 1:
                    continue
                wp = leg[pos]
                for z in zones:
                    if float(np.linalg.norm(np.asarray(wp, float) - z['c'])) < z['r'] - 1e-9 \
                            and max(t_s, z['t0']) < min(t_s + wt, z['t1']):
                        viol.append((r['veh'], 'wait', k + 1, z['zone_id']))
            t = r['dep_min'][k]
        t_arr = r['end_s'] / 60.0
        t_dep = r.get('return_dep_min', r['dep_min'][-1])
        lw = [(t_s, wt, pos) for (k_, t_s, wt, kind, pos) in r.get('waits', [])
              if k_ == len(r['seq']) + 1]
        for (u, v), (t0, t1) in zip(zip(r['legs_wp'][-1][:-1], r['legs_wp'][-1][1:]),
                                    seg_times(r['legs_wp'][-1], t_dep, lw)):
            for z in zones:
                pr = seg_cross_params(u, v, z)
                if pr is not None:
                    ta = t0 + pr[0] * (t1 - t0)
                    tb = t0 + pr[1] * (t1 - t0)
                    if max(ta, z['t0']) < min(tb, z['t1']):
                        viol.append((r['veh'], 'leg-return', 0, z['zone_id']))
        k_ret = len(r['seq']) + 1
        for (k_, t_s, wt, kind, pos) in r.get('waits', []):
            if k_ != k_ret:
                continue
            wp = r['legs_wp'][-1][pos]
            for z in zones:
                if float(np.linalg.norm(np.asarray(wp, float) - z['c'])) < z['r'] - 1e-9 \
                        and max(t_s, z['t0']) < min(t_s + wt, z['t1']):
                    viol.append((r['veh'], 'wait-return', 0, z['zone_id']))
    assert len(seen) == M, f'{case}: 覆盖 {len(seen)}/{M}'
    return viol


def export():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '表4_问题三结果'
    ws.append(['测试算例', '无人机数量N', '单架无人机最长工作时间Tmax(h)',
               '单架无人机最短工作时间Tmin(h)', 'δ=|Tmax−Tmin|(h)'])
    for case in CASES:
        with open(os.path.join(A_DIR, 'output', f'p3_{case}.json'), encoding='utf-8') as f:
            data = json.load(f)
        routes = data['routes']
        viol = verify(case, routes)
        assert not viol, f'{case} 违例: {viol[:5]}'
        s = data['summary']
        ws.append([case, s['N'], s['tmax_h'], s['tmin_h'], s['delta_h']])
        print(f'{case} 校验通过: N={s["N"]} Tmax={s["tmax_h"]}h Tmin={s["tmin_h"]}h '
              f'δ={s["delta_h"]}h 违例=0')

        p = load_points(case)
        ws2 = wb.create_sheet(f'{case}_调度方案')
        ws2.append(['无人机编号', '顺序号', '巡检点ID', '巡检等级', '第几次巡检',
                    '到达时刻(8:00起)', '离开时刻', '本段距离(单位,含绕行)',
                    '本段航路点数(含绕行点)', '本机累计飞行距离(单位)', '本机总工作时长(h)'])
        for r in routes:
            for k, s in enumerate(r['seq']):
                pid = s['point_id']
                ws2.append([r['veh'], k + 1, pid, p['level'][pid - 1], s['visit_no'],
                            hhmm(r['arr_min'][k]), hhmm(r['dep_min'][k]), r['legs_u'][k],
                            len(r['legs_wp'][k]), round(sum(r['legs_u'][:k + 1]), 2),
                            r['busy_h']])
        # 等待记录 sheet(k, t_s, wt, kind, pos;pos=0 原地 / 1 圆外悬停 / -1 到达点)
        ws3 = wb.create_sheet(f'{case}_等待与绕行')
        ws3.append(['无人机编号', '位置(段号/任务)', '类型', '等待(分钟)', '等待位置'])
        n_wait = 0
        for r in routes:
            for (k, t_s, wt, kind, pos) in r.get('waits', []):
                pos_name = ('原地(段起点)' if pos == 0 else
                            ('圆外悬停点' if pos == 1 else '到达点'))
                ws3.append([r['veh'],
                            f'段{k}' if kind.startswith('leg') else
                            f'任务{r["seq"][k-1]["point_id"] if 0 < k <= len(r["seq"]) else "?"}',
                            kind, wt, pos_name])
                n_wait += 1
        ws3.append([])
        ws3.append([f'等待事件合计', n_wait, '', '', ''])
    out = os.path.join(A_DIR, 'result3.xlsx')
    wb.save(out)
    print('\n已写入:', out)


if __name__ == '__main__':
    export()
